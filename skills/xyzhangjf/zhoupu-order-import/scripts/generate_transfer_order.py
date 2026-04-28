# -*- coding: utf-8 -*-
"""
舟谱系统调拨订单导入模板生成器

从下单表 + 价格表生成舟谱系统可导入的调拨订单Excel文件。

用法:
  python generate_transfer_order.py --order <下单表xlsx> --price <价格表xlsx> --arrival <到货日期YYYYMMDD> [--sheet <工作表名关键词>] [--start-seq <起始序号>] [--output <输出路径>]

参数:
  --order        下单表Excel文件路径（必填）
  --price        价格表Excel文件路径（必填，用于获取条码对应单位）
  --arrival      到货日期，格式YYYYMMDD（必填）
  --sheet        下单表工作表名关键词（可选，默认自动匹配到货日期）
  --start-seq    源单据号起始序号，默认1
  --output       输出文件路径（可选，默认: 调拨订单导入模板_<到货日期>.xlsx）
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import argparse
import sys

if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')
if sys.stderr.encoding and sys.stderr.encoding.lower() != 'utf-8':
    sys.stderr.reconfigure(encoding='utf-8')

# 调拨业务员列表（下单表中的列名）
TRANSFER_STAFF = ["程欢欢", "刘善涛", "毛辉", "周运潘", "田顺达", "王琴", "刘正宝", "秦小芳"]

# 固定值
TRANSFER_OUT_WAREHOUSE = "总仓"


def read_合计_row(df, staff_cols):
    """读下单表底部合计行（row.iloc[2] == '合计：'），返回 {业务员列名: 数量}"""
    for _, row in df.iterrows():
        if str(row.iloc[2]) == '合计：':
            result = {}
            for col in staff_cols:
                val = row.get(col)
                if pd.notna(val) and val > 0:
                    result[col] = int(val)
            return result
    return {}


def _simplify_name(name):
    """去掉产品简称末尾的+Nd到货备注，提取产品基础名"""
    if not name:
        return ''
    import re
    name = str(name).strip()
    name = re.sub(r'\+\d到货\（[^）]+\）', '', name)
    name = re.sub(r'\+\d发货\（[^）]+\）', '', name)
    name = re.sub(r'\（[^）]+\）$', '', name)
    return name.strip()


def load_price_table(price_file):
    """读取价格表，返回 ({条码: {单位, 简称}}, {简化名: {单位, 简称}})"""
    price_df = pd.read_excel(price_file, sheet_name='数据源')
    barcode_info = {}
    name_info = {}
    for _, row in price_df.iterrows():
        barcode_raw = row.get('条码')
        简称 = row.get('简称')
        if pd.isna(简称):
            continue
        simple_name = _simplify_name(简称)
        item = {'单位': row.get('单位'), '简称': 简称}
        if pd.notna(barcode_raw):
            barcode = str(int(barcode_raw)) if isinstance(barcode_raw, float) else str(barcode_raw).strip()
            if barcode and barcode not in barcode_info:
                barcode_info[barcode] = item
        if simple_name and simple_name not in name_info:
            name_info[simple_name] = item
    return barcode_info, name_info


def select_sheet(order_file, sheet_keyword=None, arrival_date=None):
    """选择下单表工作表"""
    xl = pd.ExcelFile(order_file)
    all_sheets = xl.sheet_names

    if sheet_keyword:
        matched = [s for s in all_sheets if sheet_keyword in s]
        if matched:
            return matched[0]
        print(f"⚠️ 未找到包含'{sheet_keyword}'的工作表，可用工作表: {all_sheets}")

    if arrival_date:
        month = str(int(arrival_date[4:6]))
        day = str(int(arrival_date[6:8]))
        for s in all_sheets:
            if month in s and day in s:
                return s

    if len(all_sheets) == 1:
        return all_sheets[0]

    print(f"⚠️ 多个工作表: {all_sheets}")
    print(f"   请用 --sheet 参数指定工作表名关键词")
    return None


def process_order_sheet(df, barcode_info, name_info):
    """处理下单表，提取调拨业务员的订单数据"""
    staff_cols = [c for c in df.columns if c in TRANSFER_STAFF]

    if not staff_cols:
        print(f"⚠️ 下单表中未找到调拨业务员列。现有列: {list(df.columns)}")
        print(f"   期望找到: {TRANSFER_STAFF}")
        return {}, []

    staff_orders = {}
    missing_barcode = []
    skipped_no_match = []

    for _, row in df.iterrows():
        barcode_raw = row.get('条码')
        product_name = row.get('简称', '')

        info = None
        barcode_str = ''

        if pd.isna(barcode_raw):
            # 无条码 → 按简称匹配
            simple_name = _simplify_name(product_name)
            info = name_info.get(simple_name)
            if info:
                barcode_str = ''
            else:
                skipped_no_match.append(product_name)
                continue
        else:
            barcode_str = str(int(barcode_raw)) if isinstance(barcode_raw, float) else str(barcode_raw).strip()
            info = barcode_info.get(barcode_str)

        for col in staff_cols:
            qty = row.get(col)
            if pd.isna(qty) or qty <= 0:
                continue

            # 调入仓 = 业务员名 + "仓"
            warehouse_in = col + "仓"

            # 获取单位
            unit = ''
            if info:
                unit = info['单位'] if pd.notna(info.get('单位')) else ''
            else:
                missing_barcode.append(barcode_str)

            if warehouse_in not in staff_orders:
                staff_orders[warehouse_in] = []
            staff_orders[warehouse_in].append({
                '条码': int(barcode_str) if barcode_str else None,
                '条码_str': barcode_str,
                '单位': str(unit) if unit else '',
                '数量': int(qty),
            })

    if skipped_no_match:
        print(f"⚠️ 无条码且无法匹配简称的行: {skipped_no_match}")
    return staff_orders, missing_barcode


def generate_excel(staff_orders, arrival_date, start_seq=1):
    """生成舟谱调拨订单导入格式的Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = '调拨订单'

    headers = [
        '*源单据号', '*调出仓', '*调入仓', '*单位', '整单备注',
        '商品编号', '商品名称', '商品条码', '期望生产日期',
        '*订单数量', '调拨参考价', '明细备注'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row_idx = 2
    seq = start_seq

    for warehouse_in in sorted(staff_orders.keys()):
        order_no = f"DB{arrival_date}{str(seq).zfill(2)}"
        seq += 1
        for item in staff_orders[warehouse_in]:
            ws.cell(row=row_idx, column=1, value=order_no)            # *源单据号
            ws.cell(row=row_idx, column=2, value=TRANSFER_OUT_WAREHOUSE)  # *调出仓
            ws.cell(row=row_idx, column=3, value=warehouse_in)          # *调入仓
            ws.cell(row=row_idx, column=4, value=item['单位'])          # *单位
            ws.cell(row=row_idx, column=8, value=item['条码'])          # 商品条码
            ws.cell(row=row_idx, column=10, value=item['数量'])         # *订单数量
            row_idx += 1

    return wb, row_idx - 2


def verify_required_fields(output_file):
    """验证所有必填字段无空值"""
    df = pd.read_excel(output_file, sheet_name='调拨订单')
    # 条码允许为空（无条码产品）
    required = ['*源单据号', '*调出仓', '*调入仓', '*单位', '*订单数量']
    all_ok = True
    issues = []
    for col in required:
        null_count = df[col].isna().sum()
        if null_count > 0:
            all_ok = False
            null_rows = df[df[col].isna()]
            for _, r in null_rows.iterrows():
                issues.append(f"  {col}为空: 调入仓={r.get('*调入仓','')}")
    # 条码空值单独提示
    barcode_null = df['商品条码'].isna().sum()
    if barcode_null > 0:
        print(f"  ⚠️ {barcode_null} 条记录条码为空（无条码产品，需手动处理）:")
    return all_ok, issues


def main():
    parser = argparse.ArgumentParser(description='舟谱系统调拨订单导入模板生成器')
    parser.add_argument('--order', required=True, help='下单表Excel文件路径')
    parser.add_argument('--price', required=True, help='价格表Excel文件路径')
    parser.add_argument('--arrival', required=True, help='到货日期YYYYMMDD')
    parser.add_argument('--sheet', default=None, help='工作表名关键词')
    parser.add_argument('--start-seq', type=int, default=1, help='源单据号起始序号（默认1）')
    parser.add_argument('--output', default=None, help='输出文件路径')
    args = parser.parse_args()

    print(f"📅 到货日期: {args.arrival}")
    print(f"📄 下单表: {args.order}")
    print(f"💰 价格表: {args.price}")

    # 1. 读取价格表（获取条码→单位映射）
    print("\n📖 读取价格表...")
    barcode_info, name_info = load_price_table(args.price)
    print(f"   共 {len(barcode_info)} 个条码, {len(name_info)} 个无条码产品")

    # 2. 选择工作表
    print("\n📖 读取下单表...")
    sheet_name = select_sheet(args.order, args.sheet, args.arrival)
    if sheet_name is None:
        sys.exit(1)
    print(f"   工作表: {sheet_name}")

    df = pd.read_excel(args.order, sheet_name=sheet_name)
    print(f"   共 {len(df)} 行")

    # ⚠️ 关键：先读合计行作为数量基准
    合计_原始 = read_合计_row(df, TRANSFER_STAFF)
    if 合计_原始:
        print(f"\n📋 下单表合计行（基准）:")
        for col, val in sorted(合计_原始.items()):
            print(f"   {col}: {val}件")
    else:
        print("\n⚠️ 未找到合计行！")

    # 3. 处理订单数据
    staff_orders, missing_barcode = process_order_sheet(df, barcode_info, name_info)

    if missing_barcode:
        print(f"\n⚠️ 以下条码在价格表中不存在:")
        for b in set(missing_barcode):
            print(f"   {b}")

    if not staff_orders:
        print("\n❌ 未生成任何调拨订单记录，请检查下单表中是否有调拨业务员的下单数据")
        sys.exit(1)

    # 4. 生成Excel
    total_records = sum(len(items) for items in staff_orders.values())
    print(f"\n📦 生成调拨订单: {len(staff_orders)} 个仓库, {total_records} 条记录")

    wb, _ = generate_excel(staff_orders, args.arrival, args.start_seq)

    output = args.output or f"调拨订单导入模板_{args.arrival}.xlsx"
    wb.save(output)
    print(f"💾 已保存: {output}")

    # ⚠️ 生成后核对：统计生成文件里每个仓库的实际件数，与合计行对比
    if 合计_原始:
        import openpyxl as xl
        gen_totals = {}
        wb_check = xl.load_workbook(output, data_only=True)
        ws_check = wb_check.active
        for row in ws_check.iter_rows(min_row=2, values_only=True):
            if row[0] is None: continue
            wh = row[2]   # 调入仓列
            qty = row[9] if row[9] else 0  # 订单数量列
            gen_totals[wh] = gen_totals.get(wh, 0) + qty

        print(f"\n🔍 生成结果 vs 下单表合计行:")
        mismatch = False
        for col, expected in sorted(合计_原始.items()):
            wh_name = col + "仓"
            actual = int(gen_totals.get(wh_name, 0))
            if actual == expected:
                print(f"   ✅ {wh_name}: {actual}件")
            else:
                print(f"   ❌ {wh_name}: 生成={actual}件 / 下单表={expected}件（差{actual-expected}）")
                mismatch = True
        for wh in gen_totals:
            if wh not in [c + "仓" for c in 合计_原始]:
                print(f"   ⚠️ 生成中有合计行无此仓库: {wh} ({int(gen_totals[wh])}件)")
        if mismatch:
            print("\n⚠️ 数量不一致，请检查后再发送！")

    # 5. 验证必填字段
    all_ok, issues = verify_required_fields(output)
    if all_ok:
        print("✅ 所有必填字段均无空值")
    else:
        print("❌ 必填字段有空值:")
        for issue in issues:
            print(issue)

    # 6. 订单统计
    print(f"\n=== 订单统计 ===")
    for name in sorted(staff_orders.keys()):
        items = staff_orders[name]
        print(f"  {name}: {len(items)}条")


if __name__ == '__main__':
    main()
