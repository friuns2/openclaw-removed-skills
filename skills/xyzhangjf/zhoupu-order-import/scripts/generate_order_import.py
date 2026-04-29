# -*- coding: utf-8 -*-
"""
舟谱系统自提订单导入模板生成器 v2（2026-04-09 更新）

从下单表 + 价格表生成舟谱系统可导入的自提订单Excel文件。

核心规则（2026-04-09）：
- 一个客户只能有一个源单据号（该客户所有商品行共用同一单号）
- 模板字段（9列）：源单据号、客户名称、业务员、部门、仓库、商品名称、商品条码、单位、数量
- 分销商（唐成/黄家伟/易胜琳→易胜玲/胡魁魁/朱青峰/谢总→宜城谢总）+ 门店（吾悦/东津/民发→永辉；沃尔玛；檀溪美联→美联）都填入自提订单
- 业务员=张俊峰，部门=湖北福宝商贸有限公司，仓库=总仓

用法:
  python generate_order_import.py --order <下单表xlsx> --price <价格表xlsx> --arrival <YYYYMMDD>
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import argparse, json, sys, os, re

if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')
if sys.stderr.encoding and sys.stderr.encoding.lower() != 'utf-8':
    sys.stderr.reconfigure(encoding='utf-8')

# ============================================================
# 内置客户配置
# ============================================================
DEFAULT_CONFIG = {
    # 下单表列名 → 客户类型（分销商+导购门店 = 全部自提订单客户）
    "distributors": ["唐成", "黄家伟", "易胜琳", "胡奎奎", "朱青峰", "谢总"],
    "stores":       ["吾悦", "东津", "民发", "沃尔玛", "檀溪美联"],

    # 下单表列名 → 舟谱系统客户名称（注意：下单表列名要精确匹配！）
    "customer_mapping": {
        "易胜琳":  "易胜玲",   # 易胜琳（下单表）→ 易胜玲（舟谱）
        "谢总":    "宜城谢总",  # 谢总（下单表）→ 宜城谢总（舟谱）
        "吾悦":    "永辉吾悦店",
        "东津":    "永辉东津店",
        "民发":    "永辉民发店",
        "檀溪美联": "美联檀溪店",
    },

    # 下单表列名 → 价格表价格列名
    "price_rules": {
        "唐成":    "分销价格",
        "黄家伟":  "分销价格",
        "易胜琳":  "分销价格",
        "胡奎奎":  "分销价格",
        "朱青峰":  "分销价格",
        "谢总":    "分销价格",
        "吾悦":    "永辉价格",
        "东津":    "永辉价格",
        "民发":    "永辉价格",
        "沃尔玛":  "沃尔玛价格",
        "檀溪美联": "美联价格",
    },

    # 固定字段
    "业务员": "张俊峰",
    "部门":   "湖北福宝商贸有限公司",
    "仓库":   "总仓",
}


def load_config(path=None):
    cfg = json.loads(json.dumps(DEFAULT_CONFIG))
    if path and os.path.exists(path):
        with open(path, 'r', encoding='utf-8') as f:
            cfg.update(json.load(f))
    return cfg


def _simplify(name):
    """去掉产品名称末尾的+Nd到货/发货备注"""
    if not name: return ''
    s = str(name).strip()
    s = re.sub(r'\+\d到货[（(][^）)]*[）)]', '', s)
    s = re.sub(r'\+\d发货[（(][^）)]*[）)]', '', s)
    s = re.sub(r'[（(][^）)]*[）)]$', '', s)
    return s.strip()


def _strip_unit(name):
    """去掉名称中的规格单位部分，如125g*2杯、250ml、200g*10瓶*6组等，保留核心品名"""
    s = str(name).strip()
    # 按常见单位/规格模式截断
    s = re.sub(r'\d+g\*?\d*杯.*$', '', s)       # 125g*2杯、200g杯
    s = re.sub(r'\d+ml.*$', '', s)               # 250ml
    s = re.sub(r'\d+g.*$', '', s)               # 195g
    s = re.sub(r'\*?\d+瓶.*$', '', s)           # *10瓶*6组
    s = re.sub(r'\*?\d+组.*$', '', s)           # *6组
    s = re.sub(r'\*\d+.*$', '', s)              # *10
    return s.strip()


def _find_by_name(nm, name):
    """智能按名称匹配：精确→简化→品名前缀匹配（优先匹配带条码的最短key）"""
    if not name: return None
    # 1. 精确
    if name in nm: return nm[name]
    # 2. 简化（去+Nd备注）
    simple = _simplify(name)
    if simple in nm: return nm[simple]
    # 3. 品名前缀匹配：nm的key是simple的前缀（key ⊂ simple）
    #    优先选有条码的、key最短的（最精确匹配）
    best, best_key_len = None, 999
    for key, info in nm.items():
        if simple.startswith(key):
            has_bc = bool(info.get('条码'))
            if has_bc and (best is None or not best.get('条码') or len(key) < best_key_len):
                best = info
                best_key_len = len(key)
            elif best is None:
                best = info
                best_key_len = len(key)
    return best


def load_price_table(path):
    """返回 ({条码: {单位,简称,...}}, {简化名: {单位,简称,...}})"""
    df = pd.read_excel(path, sheet_name='数据源')
    barcode_map = {}
    name_map = {}   # 简化名 → 产品信息（处理无条码产品）

    for _, row in df.iterrows():
        简称 = row.get('简称')
        if pd.isna(简称): continue

        item = {
            '单位':          str(row.get('单位', '')) if pd.notna(row.get('单位')) else '',
            '简称':          简称,
            '分销价格':      _round(row.get('分销价格')),
            '永辉价格':      _round(row.get('永辉价格')),
            '沃尔玛价格':    _round(row.get('沃尔玛价格')),
            '美联价格':      _round(row.get('美联价格')),
        }

        bc_raw = row.get('条码')
        if pd.notna(bc_raw):
            bc = str(int(bc_raw)) if isinstance(bc_raw, float) else str(bc_raw).strip()
            if bc:
                if bc not in barcode_map:
                    barcode_map[bc] = item
                else:
                    # 补充空值
                    for k in ['分销价格','永辉价格','沃尔玛价格','美联价格']:
                        if barcode_map[bc][k] is None:
                            barcode_map[bc][k] = item[k]

        # 无条码产品也建简称映射
        simple = _simplify(简称)
        if simple and simple not in name_map:
            name_map[simple] = item

    return barcode_map, name_map


def _round(v):
    if pd.isna(v): return None
    return round(float(v), 2)


def apply_extra_prices(bm, nm, extra):
    """补充价格表缺失的条码价格"""
    for bc, data in extra.items():
        bc = str(bc).strip()
        if bc not in bm:
            bm[bc] = {'单位': data.get('单位',''), '简称': data.get('简称',''),
                      '分销价格':None,'永辉价格':None,'沃尔玛价格':None,'美联价格':None}
        for k in ['分销价格','永辉价格','沃尔玛价格','美联价格']:
            if k in data and data[k] is not None:
                bm[bc][k] = round(float(data[k]), 2)
        for k in ['单位','简称']:
            if k in data and data[k]:
                bm[bc][k] = data[k]
        # 用简称更新 nm
        简称 = data.get('简称','')
        if 简称:
            # 遍历nm所有key，对匹配到的条目补充条码和价格
            simple = _simplify(简称)
            updated = False
            for existing_key in list(nm.keys()):
                if existing_key.startswith(simple) or simple.startswith(existing_key):
                    # 这个nm条目对应同一个产品，补充条码
                    if not nm[existing_key].get('条码'):
                        nm[existing_key] = dict(nm[existing_key])
                        nm[existing_key]['条码'] = bc
                    # 补充价格
                    for k in ['分销价格','永辉价格','沃尔玛价格','美联价格']:
                        if k in data and data[k] is not None and not nm[existing_key].get(k):
                            nm[existing_key][k] = round(float(data[k]), 2)
                    updated = True
            # 如果没有匹配到的key，新增
            if not updated:
                bm[bc]['条码'] = bc
                nm[simple] = bm[bc].copy()
    return bm, nm


def select_sheet(order_file, keyword=None, arrival=None):
    xl = pd.ExcelFile(order_file)
    sheets = xl.sheet_names
    if keyword:
        matched = [s for s in sheets if keyword in s]
        if matched: return matched[0]
        print(f"⚠️ 未找到含'{keyword}'的sheet，可用: {sheets}")
    if arrival:
        m = str(int(arrival[4:6]))
        d = str(int(arrival[6:8]))
        for s in sheets:
            if m in s and d in s: return s
    if len(sheets) == 1: return sheets[0]
    print(f"⚠️ 多个工作表: {sheets}，请用 --sheet 指定")
    return None


def read_合计_row(df, customer_cols):
    """读下单表底部合计行（row[2] == '合计：'），返回 {下单表列名: 数量}

    ⚠️ 重要：永远用这个作为数量基准，不要自己累加个别行。
    原因：个别行含公式导致data_only读值不准；中间嵌有"加单汇总行"会重复计算。
    """
    for _, row in df.iterrows():
        if str(row.iloc[2]) == '合计：':
            result = {}
            for col in customer_cols:
                val = row.get(col)
                if pd.notna(val) and val > 0:
                    result[col] = int(val)
            return result
    return {}


def process_sheet(df, config, bm, nm):
    """处理下单表，返回 [{客户名, 订单行[{条码,条码str,品名,单位,数量,单价}]}]"""
    all_cols = config['distributors'] + config['stores']
    customer_cols = [c for c in df.columns if c in all_cols]
    if not customer_cols:
        print(f"⚠️ 未找到客户列。下单表列: {list(df.columns)}")
        return {}, [], []

    customer_orders = {}
    missing_price, skipped = [], []

    for _, row in df.iterrows():
        # 跳过合计行（条码和名称都为空的行）
        bc_raw = row.get('条码')
        简称 = row.get('简称', '')
        if pd.isna(bc_raw) and (pd.isna(简称) or str(简称).strip() == ''):
            continue

        # 定位产品信息
        info, bc_str = None, ''
        if pd.isna(bc_raw):
            # 无条码 → 用智能名称匹配（name_map里已存条码）
            info = _find_by_name(nm, str(简称) if pd.notna(简称) else '')
            if info:
                bc_str = str(info.get('条码', '')) or ''
                if not bc_str:
                    print(f'  ⚠️ 匹配成功但无条码: 简称="{简称}" → info简称="{info.get("简称","")}"')
        else:
            bc_str = str(int(bc_raw)) if isinstance(bc_raw, float) else str(bc_raw).strip()
            info = bm.get(bc_str)
            # 有条码但价格表无此条码 → 按简称兜底
            if not info:
                info = _find_by_name(nm, str(简称) if pd.notna(简称) else '')

        if not info:
            skipped.append(str(简称))
            continue

        # 商品名称：优先用价格表简称，没有则用下单表简称
        product_name = info.get('简称', '') if info else ''
        if not product_name:
            product_name = str(简称) if pd.notna(简称) else ''

        for col in customer_cols:
            qty = row.get(col)
            if pd.isna(qty) or qty <= 0: continue

            cust_name = config['customer_mapping'].get(col, col)
            price_col = config['price_rules'].get(col, '分销价格')
            price = info.get(price_col)
            if price is None:
                missing_price.append(f"{bc_str} {product_name} → {cust_name}无{price_col}")

            if cust_name not in customer_orders:
                customer_orders[cust_name] = []
            customer_orders[cust_name].append({
                '条码':     int(bc_str) if bc_str else None,
                '条码_str': bc_str,
                '品名':     product_name,
                '单位':     info.get('单位', ''),
                '数量':     int(qty),
                '单价':     float(price) if price else None,
            })

    return customer_orders, missing_price, skipped


def generate_excel(customer_orders, config, arrival_date, start_seq=1):
    """生成9列自提订单导入模板：一个客户共用一个源单据号"""
    wb = Workbook()
    ws = wb.active
    ws.title = '自提订单'

    # 19列完整模板（参考正确版）
    headers = [
        '*源单据号',      # 0
        '客户编号',       # 1  （留空）
        '客户名称',       # 2
        '*业务员',        # 3
        '部门',           # 4
        '*仓库',          # 5
        '单据日期',       # 6  （留空）
        '整单备注',       # 7  （留空）
        '制单人',         # 8  （留空）
        '商品编号',       # 9  （留空）
        '商品货号',       # 10 （留空）
        '商品名称',       # 11 （留空，舟谱靠条码识别商品）
        '商品条码',       # 12
        '*单位',          # 13
        '*数量',          # 14
        '*单价(折后价)',  # 15
        '明细备注',       # 16 （留空）
        '业务属性',       # 17 （留空）
        '标签',           # 18 （留空）
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row_idx = 2
    seq = start_seq

    for cust in sorted(customer_orders.keys()):
        # 一个客户 → 一个源单据号
        order_no = f"ZT{arrival_date}{str(seq).zfill(4)}"
        seq += 1
        for item in customer_orders[cust]:
            ws.cell(row=row_idx, column=1,  value=order_no)           # *源单据号
            ws.cell(row=row_idx, column=2,  value=None)                # 客户编号（空）
            ws.cell(row=row_idx, column=3,  value=cust)                # 客户名称（舟谱系统名）
            ws.cell(row=row_idx, column=4,  value=config['业务员'])   # *业务员
            ws.cell(row=row_idx, column=5,  value=config['部门'])      # 部门
            ws.cell(row=row_idx, column=6,  value=config['仓库'])      # *仓库
            ws.cell(row=row_idx, column=7,  value=None)                # 单据日期（空）
            ws.cell(row=row_idx, column=8,  value=None)                # 整单备注（空）
            ws.cell(row=row_idx, column=9,  value=None)                # 制单人（空）
            ws.cell(row=row_idx, column=10, value=None)                # 商品编号（空）
            ws.cell(row=row_idx, column=11, value=None)                # 商品货号（空）
            ws.cell(row=row_idx, column=12, value=None)                # 商品名称（空，舟谱靠条码识别）
            ws.cell(row=row_idx, column=13, value=item['条码'])        # 商品条码
            ws.cell(row=row_idx, column=14, value=item['单位'])         # *单位
            ws.cell(row=row_idx, column=15, value=item['数量'])        # *数量
            ws.cell(row=row_idx, column=16, value=item['单价'])         # *单价(折后价)
            ws.cell(row=row_idx, column=17, value=None)                # 明细备注（空）
            ws.cell(row=row_idx, column=18, value=None)                # 业务属性（空）
            ws.cell(row=row_idx, column=19, value=None)                # 标签（空）
            row_idx += 1

    return wb, row_idx - 2


def verify(output_file):
    """验证必填字段"""
    df = pd.read_excel(output_file, sheet_name='自提订单')
    # 必填（*标记）：0=源单据号, 3=业务员, 5=仓库, 12=商品条码, 13=单位, 14=数量, 15=单价
    header = list(df.columns)
    required_names = ['*源单据号', '*业务员', '*仓库', '商品条码', '*单位', '*数量', '*单价(折后价)']
    all_ok, issues = True, []
    for col_name in required_names:
        if col_name not in header: continue
        nulls = df[df[col_name].isna()]
        if len(nulls) > 0:
            all_ok = False
            for _, r in nulls.iterrows():
                issues.append(f"  {col_name} 为空: 客户={r.get('客户名称','')} 条码={r.get('商品条码','')}")
    return all_ok, issues


def main():
    parser = argparse.ArgumentParser(description='舟谱自提订单导入模板生成器 v2')
    parser.add_argument('--order',            required=True)
    parser.add_argument('--price',            required=True)
    parser.add_argument('--arrival',          required=True, help='到货日期YYYYMMDD')
    parser.add_argument('--sheet',            default=None)
    parser.add_argument('--start-seq',        type=int, default=1)
    parser.add_argument('--output',           default=None)
    parser.add_argument('--extra-prices',     default=None)
    parser.add_argument('--extra-prices-file',default=None)
    parser.add_argument('--config',           default=None)
    args = parser.parse_args()

    cfg = load_config(args.config)

    print(f"📅 到货日期: {args.arrival}")
    print(f"📄 下单表: {args.order}")
    print(f"💰 价格表: {args.price}")

    # 价格表
    print("\n📖 读取价格表...")
    bm, nm = load_price_table(args.price)
    print(f"   {len(bm)} 个条码, {len(nm)} 个无条码产品")

    # 补充价格
    extra = {}
    if args.extra_prices:
        extra.update(json.loads(args.extra_prices))
    if args.extra_prices_file and os.path.exists(args.extra_prices_file):
        with open(args.extra_prices_file, 'r', encoding='utf-8') as f:
            extra.update(json.load(f))
    if extra:
        bm, nm = apply_extra_prices(bm, nm, extra)
        print(f"   补充 {len(extra)} 个条码价格")

    # 下单表
    print("\n📖 读取下单表...")
    sheet = select_sheet(args.order, args.sheet, args.arrival)
    if not sheet: sys.exit(1)
    print(f"   工作表: {sheet}")
    df = pd.read_excel(args.order, sheet_name=sheet)
    print(f"   共 {len(df)} 行")

    # ⚠️ 关键：先读合计行作为数量基准
    all_customer_cols = cfg['distributors'] + cfg['stores']
    customer_cols_in_df = [c for c in df.columns if c in all_customer_cols]
    合计_原始 = read_合计_row(df, customer_cols_in_df)
    if 合计_原始:
        print(f"\n📋 下单表合计行（基准）:")
        for col, val in sorted(合计_原始.items()):
            mapped = cfg['customer_mapping'].get(col, col)
            print(f"   {col}（{mapped}）: {val}件")
    else:
        print("\n⚠️ 未找到合计行！")

    # 处理
    orders, miss_price, skipped = process_sheet(df, cfg, bm, nm)

    if skipped:
        print(f"\n⚠️ 无法匹配的行（无条码且简称未找到）: {skipped}")

    if miss_price:
        print(f"\n⚠️ 缺少价格:")
        for m in miss_price: print(f"   {m}")
        print("\n补充格式: --extra-prices-file <JSON文件>")
        print('  JSON格式: {"条码": {"分销价格": 10.5, "永辉价格": 12.0}}')

    if not orders:
        print("\n❌ 未生成任何订单")
        sys.exit(1)

    # 统计
    total = sum(len(v) for v in orders.values())
    print(f"\n📦 {len(orders)} 个客户, {total} 条记录")
    for name in sorted(orders.keys()):
        items = orders[name]
        null_price = sum(1 for i in items if i['单价'] is None)
        print(f"  {name}: {len(items)}条" + (f" ⚠️{null_price}条缺单价" if null_price else ""))

    # 生成
    wb, rec_cnt = generate_excel(orders, cfg, args.arrival, args.start_seq)
    out = args.output or f"自提订单导入模板_{args.arrival}.xlsx"
    wb.save(out)
    print(f"\n💾 已保存: {out}")

    # ⚠️ 生成后核对：统计生成文件里每个客户的实际件数，与合计行对比
    if 合计_原始:
        import openpyxl as xl
        gen_totals = {}
        wb_check = xl.load_workbook(out, data_only=True)
        ws_check = wb_check.active
        for row in ws_check.iter_rows(min_row=2, values_only=True):
            if row[0] is None: continue
            cust = row[2]   # 客户名称列（舟谱系统名）
            qty  = row[14] if row[14] else 0  # 数量列
            gen_totals[cust] = gen_totals.get(cust, 0) + qty

        print(f"\n🔍 生成结果 vs 下单表合计行:")
        mismatch = False
        for col, expected in sorted(合计_原始.items()):
            mapped = cfg['customer_mapping'].get(col, col)
            actual = int(gen_totals.get(mapped, 0))
            if actual == expected:
                print(f"   ✅ {mapped}: {actual}件")
            else:
                print(f"   ❌ {mapped}: 生成={actual}件 / 下单表={expected}件（差{actual-expected}）")
                mismatch = True
        for cust in gen_totals:
            if cust not in [cfg['customer_mapping'].get(col, col) for col in 合计_原始]:
                print(f"   ⚠️ 生成中有合计行无此客户: {cust} ({int(gen_totals[cust])}件)")
        if mismatch:
            print("\n⚠️ 数量不一致，请检查后再发送！")

    # 验证
    ok, issues = verify(out)
    if ok:
        print("✅ 所有必填字段均无空值")
    else:
        print("❌ 必填字段有空值:")
        for i in issues: print(i)


if __name__ == '__main__':
    main()
